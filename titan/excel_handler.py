"""
Module for MS Excel file handling.

:author: Pekka Savolainen <pekka.t.savolainen@vtt.fi>
:date:   25.8.2016
"""

import logging
import os
import openpyxl
from openpyxl import load_workbook


class ExcelHandler:
    def __init__(self, path):
        """Class constructor."""
        self.path = path
        self.wb = None

    def load_wb(self):
        """Load Excel workbook."""
        logging.debug("Opening Excel Workbook {0}".format(self.path))
        try:
            self.wb = load_workbook(filename=self.path)
        except:
            raise

    def sheet_names(self):
        """Return a list of sheet names in the opened workbook."""
        return self.wb.get_sheet_names()

    def read_project_sheet(self):
        """Read Project details from Project sheet."""
        try:
            project_sheet = self.wb.get_sheet_by_name('Project')
        except KeyError:
            logging.error("Project sheet not found")
            return False
        name = project_sheet['B1'].value
        desc = project_sheet['B2'].value
        work_dir = project_sheet['B3'].value
        return [name, desc, work_dir]

    def import_setups(self):
        """Read Setups sheet."""
        setup_sheet = self.wb['Setups']
        rows = str(setup_sheet.max_row)
        columns = setup_sheet.max_column
        logging.debug("Setups sheet has {0} rows and {1} columns".format(rows, columns))
        parents = [v[0].value for v in setup_sheet['A2':'A' + rows]]
        currents = [v[0].value for v in setup_sheet['B2':'B' + rows]]
        tools = [v[0].value for v in setup_sheet['C2':'C' + rows]]
        tool_args = [v[0].value for v in setup_sheet['D2':'D' + rows]]
        descriptions = [v[0].value for v in setup_sheet['E2':'E' + rows]]
        ready = [v[0].value for v in setup_sheet['F2':'F' + rows]]
        failed = [v[0].value for v in setup_sheet['G2':'G' + rows]]
        return [parents, currents, tools, tool_args, descriptions, ready, failed]

    def read_data_sheet(self, sheet_name):
        """Read data from the given sheet.

        Args:
            sheet_name (str): Name of sheet with data

        Returns:
            List of lists including filename, Setups and other data
        """
        sheet = self.wb.get_sheet_by_name(sheet_name)
        n_rows = sheet.max_row
        n_columns = sheet.max_column
        logging.debug("Processing {0}. Includes {1} rows and {2} columns".format(sheet, n_rows, n_columns))
        # Get all data on the sheet
        data = list(sheet.rows)
        # Get header row
        header = data.pop(0)
        if n_columns == 1 and n_rows == 1 and not header[0].value:
            logging.debug("No data found on {0}".format(sheet))
            return []
        if n_columns < 4:
            logging.error("Sheet should have at least 4 columns: Filename, Setup, Set, and Value")
            return []
        headers = [item.value for item in header]  # Append header values into list
        # Number of Set columns
        n_sets = n_columns - 3  # n_columns - filename column - setup column - value column
        sets = list()
        # All data
        filename = [v[0].value for v in data]
        setup = [v[1].value for v in data]
        # Get all sets into one list
        for i in range(n_sets):
            set_i = [v[i+2].value for v in data]
            sets.append(set_i)
        value = [v[-1].value for v in data]
        return [headers, filename, setup, sets, value, n_rows]

    def export_to_excel(self):
        """Exports the selected data to the defined excel file.
        Example on how to write to Excel file.
        """
        python_date_format = '%Y-%m-%d %H:%M:%S'
        excel_date_format = 'yyyy-mm-dd hh:mm:ss'
        logging.debug('Exporting %s data' % self._title)
        # Check if directory exists and create if not
        file_dir = os.path.dirname(config.EXPORTED_DATA_FILE)
        if not os.path.exists(file_dir):
            logging.debug("Creating folder: %s" % file_dir)
            os.makedirs(file_dir)

        # Open new file or load existing
        if os.path.exists(config.EXPORTED_DATA_FILE):
            logging.debug('Opening file %s' % config.EXPORTED_DATA_FILE)
            wbook = openpyxl.load_workbook(filename=config.EXPORTED_DATA_FILE)
        else:
            logging.debug('Creating file %s' % config.EXPORTED_DATA_FILE)
            wbook = openpyxl.Workbook()

        # If the sheet already exists, delete it in order to replace it
        try:
            wbook.remove_sheet(wbook.get_sheet_by_name(self._title))
        except (KeyError, ValueError):
            pass

        # Create sheet, write title, and export timestamp
        wsheet = wbook.create_sheet()
        wsheet.title = self._title
        wsheet.cell(None, 1, 1).value = self._title
        export_stamp = 'Exported: ' \
                       + datetime.now().strftime(python_date_format)
        wsheet.cell(None, 2, 1).value = export_stamp

        sorted_scenarios = self.sort_scenarios()
        # Write the headers (configuration, scenario, date and label)
        i = 2
        for scenario in sorted_scenarios:
            conf_name = self.get_conf_name(scenario)
            wsheet.cell(None, 1, i).value = conf_name
            wsheet.cell(None, 2, i).value = scenario.scenario
            wsheet.cell(None, 3, i).value = scenario.created
            wsheet.cell(None, 3, i).number_format = excel_date_format
            wsheet.cell(None, 4, i).value = scenario.label
            i += 1

        # Writes results for time plots and fill plots.
        if self._plot_type is PlotType.TIME_PLOT or self._plot_type is PlotType.FILL_PLOT:
            all_timestamps, output_lines = self.get_timestamps()
            i = 2
            # Writes the values for each scenario
            for scenario in sorted_scenarios:
                values = self.get_data_time_plots(scenario, all_timestamps)
                for j in range(len(values)):
                    wsheet.cell(None, j+5, i).value = values[j]
                i += 1
            # Writes the timestamps in the first row
            i = 5
            for line in output_lines:
                wsheet.cell(None, i, 1).value \
                    = datetime.strptime(line[:-1], python_date_format)
                wsheet.cell(None, i, 1).number_format = excel_date_format
                i += 1

        # Delete default sheet and save file
        try:
            wbook.remove_sheet(wbook.get_sheet_by_name('Sheet'))
        except (KeyError, ValueError):
            pass
        # Handle error when file is being used
        if not self.save_and_close_excel():
            logging.debug("File not exported.")

    def save_and_close_excel(self):
        """ Handles the saving of the exported file

        Args:
            wbook: workbook being written on.

        Return:
            boolean flagging the success of the save.
        """

        try:
            self.wb.save(self.path)
        except PermissionError:
            logging.exception("Permission error. File is probably reserved for another process.")
            return False
        return True

